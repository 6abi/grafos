
#imports
import os
import sys
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
import random
from matplotlib.pyplot import figure
import numpy as np
import json
from openpyxl import load_workbook
from unicodedata import normalize
import warnings
warnings.filterwarnings("ignore")

dir = "planilhas/"

def get_file_name(filepath):
    """Lista os arquivos da pasta"""
    f = os.listdir(filepath)
    files = [os.path.splitext(filename)[0] for filename in os.listdir(dir)]
    return files



def get_sheet_names_xlsx(filepath):
    """Lista a sheet do arquivo"""
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames



def get_all_files():
    """Loop para listar todos os arquivos e suas sheets
def get_all_files():"""
    f = np.array(get_file_name(dir))
    sheets = []
    file_name = []
    for i in range(len(f)):
        file_name.append(f.item(i))
        s = get_sheet_names_xlsx(dir+file_name[i]+".xlsx")
        s_as_string = ' '.join([str(elem) for elem in s])   #converte a lista para string
        sheets.append(s_as_string)                          #nova lista
    # print("Files: {}".format(file_name))
    # print("Sheets: {}".format(sheets))
    return convert_all_to_json(dir, file_name, sheets, ".xlsx")



def convert_to_json(dir, file_name, sheet_name, ext):
    """Converte excel para Json"""
    excel_data_df = pd.read_excel(dir + file_name + ext, sheet_name=sheet_name)
    json_str = excel_data_df.to_json(orient='records')
    #print("The original dictionary is : {}".format(json_str))
    return handle_accentuation_json(json_str)



def handle_accentuation_json(json_str):
    """Tratamento de acentuação no Json"""
    json_str = json_str.encode("latin1").decode("unicode_escape")
    json_result = normalize('NFKD', json_str).encode('ASCII', 'ignore').decode('ASCII')
    #print(json_result)
    return json_result


def convert_all_to_json(dir, f, sheet_name, ext):
    """Loop para converter para Json"""
    list_json = []
    for i in range(len(f)):
        parsed =  json.loads(convert_to_json(dir, str(f[i]), sheet_name[i], ext))
        json_data = (json.dumps(parsed, indent=4, sort_keys=False))
        list_json.append(parsed)
    # printing Structed dictionary
    # print("The structed dictionary is : {}".format(list_json))
    return save_json(list_json)


def save_json(json_data):
    """Salva a estrutura JSON criada"""
    file_name = 'output/data.json'
    with open(file_name, 'w') as outfile:
        json.dump(json_data, outfile, indent=4)
    return convert_json_to_pandasDF(json_data)


def convert_json_to_pandasDF(json_data):
    """Converte o Json em um dataframe"""
    json_list = [item for sublist in json_data for item in sublist]
    dataset = pd.json_normalize(json_list)
    dataset = pd.DataFrame(dataset)
    pd.set_option("max_rows", None)
    # print(dataset)
    return cleansing_data(dataset)


def cleansing_data(df):
    """Limpeza e tratamento de dados"""
    df = df[['Fonte', 'Conexao', 'Velocidade ( Giga )']]
    df.rename(columns={'Velocidade ( Giga )': 'Giga'}, inplace = True)
    df.dropna(subset=['Conexao'], inplace=True)
    df.fillna(0, inplace=True)
    df= df[df['Conexao'] != "None"]
    # print(df.head(10))
    # print(df.shape)
    # print("Velocidades:{}".format(df['Giga'].unique()))
    # print(df['Giga'].describe())
    return graph_generator(df)


def graph_generator(df):
    """Geração do grafo """
    df['Giga'] = df['Giga'].values
    G =nx.from_pandas_edgelist(df, 'Fonte', 'Conexao', 'Giga')
    labels = nx.get_edge_attributes(G,'Giga')
    pos=nx.shell_layout(G)# posição do texto nas arestas
    nodes = G.nodes()
    sizes = []

    #grau de centralidade
    centrality = nx.degree_centrality(G)

    #tamanho dos nós
    for v in G.nodes():
        centrality[v] = centrality[v]
        # print(f"{v:2} {centrality[v]:.4f}")
        x = centrality[v] * 30000
        sizes.append(round(x,4))

    switcher = {
        1.0:  df['Giga'].unique()[0],
        10.0: df['Giga'].unique()[1]
    }
    edge_value = []
    for label in labels:
        try:
            edge_value.append(switcher[labels[label]])
        except KeyError:
            print("Desconhecido")
    # print(edge_value)

    #alterar cor aresta
    edge_colors = ['blue' if e == switcher[1] else 'red' for e in edge_value]
    # print(edge_colors)

    #configs do grafo
    options = {
        'node_color':'c',
        'edge_color':edge_colors,
        "edge_cmap": plt.cm.Blues,
        'with_labels': True,
        'node_size':sizes,
        'alpha': .95,
    }

    #grafo
    fig = figure(figsize=(20,20))
    nx.draw_shell(G, **options)
    nx.draw_networkx_edge_labels(G, pos=pos, edge_labels=labels)
    nx.draw


    #Concatenar a centralidade com o Dataframe original
    df2 = pd.DataFrame((list(centrality.items())),  columns=['Conexao', 'Grau de centralidade'])
    df2['Grau de centralidade'] = df2['Grau de centralidade'].apply(lambda x: round(x, 4))
    df = pd.merge(df,df2, how='inner', on='Conexao')
    df.rename({'Grau de centralidade': 'Grau de centralidade(Coluna Conexao)'}, axis=1, inplace=True)
    print(df)
    return save_files(fig, df, df2)


def save_files(fig, df,df2):
    fig.savefig('output/grafo2.png')
    df.to_csv('output/data_switches.csv', index=False, sep=';')
    df2.to_csv('output/centrality.csv', index=False, sep=';')


#chamada funções
def main():
    get_all_files()
    sys.exit()

#main
if __name__ == '__main__':
    main()

